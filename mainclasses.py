import customtkinter as c
import tkinter as t
from tkinter import ttk
from openpyxl import load_workbook
from datetime import datetime

class App(c.CTk):
    def __init__(self):
        super().__init__()
        self.geometry("1000x500")
        self.title("Book Inventory")
        self.iconbitmap("book.ico")
        
        # load individual frames
        self.intro = Intro(self)
        self.add = Add_book(self)
        self.bookOut = Book_Out(self)
        self.bookIn = Book_In(self)
        
        # run the app
        self.mainloop()
    

    
    
        

class Intro(c.CTkFrame):
    # this object is the frame as stated in the argument, so no need for intro_frame = c.CTkFrame(self)
    def __init__(self, parent):
        super().__init__(parent)

        # fix frame size
        self.place(x=0,y=0,relwidth=1,relheight=.1)

        # add widgets to app
        intro = c.CTkLabel(self, text="Welcome to book inventory manager",bg_color='white')
        intro.pack(expand=True, fill='both')

class Add_book(c.CTkFrame):
    # this object is the frame as stated in the argument, so no need for content_frame = c.CTkFrame(self)
    def __init__(self, parent):
        super().__init__(parent)   

        # fix frame size
        self.place(x=0,rely=.1,relwidth=1,relheight=.2)

        self.grid_columnconfigure((0,2), weight=1)
        
        # adding widgets
        self.bookIdLabel = c.CTkLabel(self, text="Enter Barcode of Book")
        self.bookIdLabel.grid(row=0, column=0)
        self.bookIdValue = c.CTkEntry(self)
        self.bookIdValue.grid(row=0, column=1)
        self.bookNameLabel = c.CTkLabel(self, text="Enter name of Book")
        self.bookNameLabel.grid(row=1, column=0)
        self.bookNameValue = c.CTkEntry(self)
        self.bookNameValue.grid(row=1, column=1)

        # adding button to submit
        self.button = c.CTkButton(self, command=self.add_book, text="Add Book")
        self.button.grid(row=2,columnspan=3)

    # add method to frame
    def add_book(self):
        
        output_file_name = 'books.xlsx'
        wb = load_workbook(output_file_name, data_only=True)
        ws = wb['Library']
        id = 1
        ws.append([int(self.bookIdValue.get()), self.bookNameValue.get()])
        wb.save(output_file_name)

class Book_Out(c.CTkFrame):
    # this object is the frame as stated in the argument, so no need for content_frame = c.CTkFrame(self)
    def __init__(self, parent):
        super().__init__(parent)   

        # fix frame size
        self.place(x=0,rely=.3,relwidth=1,relheight=.35)

        self.grid_columnconfigure((0,2), weight=1)

        # adding widgets
        self.bookIdLabel = c.CTkLabel(self, text="Enter Barcode of Book")
        self.bookIdLabel.grid(row=0, column=0)
        self.bookIdValue = c.CTkEntry(self)
        self.bookIdValue.grid(row=0, column=1)
        self.bookNameLabel = c.CTkLabel(self, text="Enter name of Book")
        self.bookNameLabel.grid(row=1, column=0)
        self.bookNameValue = c.CTkEntry(self)
        self.bookNameValue.grid(row=1, column=1)
        self.studentNameLabel = c.CTkLabel(self, text="Enter name of Student")
        self.studentNameLabel.grid(row=2, column=0)
        self.studentNameValue = c.CTkEntry(self)
        self.studentNameValue.grid(row=2, column=1)

        # adding button to submit
        self.button = c.CTkButton(self, command=self.book_out, text="Book Out")
        self.button.grid(row=3,columnspan=3)

    # add method to frame
    def book_out(self):
    
        output_file_name = 'books.xlsx'
        wb = load_workbook(output_file_name, data_only=True)

        # adding ID, book name and student name to worksheet
        ws = wb['BooksOut']
        ws.append([int(self.bookIdValue.get()), self.bookNameValue.get(),self.studentNameValue.get()])

        # adding ID, book name, student name, type of transaction and date to worksheet
        ws2 = wb['Log']
        ws2.append([int(self.bookIdValue.get()), self.bookNameValue.get(),self.studentNameValue.get(),'Out',datetime.now()])
        wb.save(output_file_name)

class Book_In(c.CTkFrame):
    # this object is the frame as stated in the argument, so no need for content_frame = c.CTkFrame(self)
    def __init__(self, parent):
        super().__init__(parent)   

        # fix frame size
        self.place(x=0,rely=.65,relwidth=1,relheight=1)

        self.grid_columnconfigure((0,2), weight=1)

        # adding widgets
        self.bookIdLabel = c.CTkLabel(self, text="Enter Barcode of Book")
        self.bookIdLabel.grid(row=0, column=0)
        self.bookIdValue = c.CTkEntry(self)
        self.bookIdValue.grid(row=0, column=1)
        self.bookNameLabel = c.CTkLabel(self, text="Enter name of Book")
        self.bookNameLabel.grid(row=1, column=0)
        self.bookNameValue = c.CTkEntry(self)
        self.bookNameValue.grid(row=1, column=1)
        self.studentNameLabel = c.CTkLabel(self, text="Enter name of Student")
        self.studentNameLabel.grid(row=2, column=0)
        self.studentNameValue = c.CTkEntry(self)
        self.studentNameValue.grid(row=2, column=1)

        # adding button to submit
        self.button = c.CTkButton(self, command=self.returnBook, text="Return Book")
        self.button.grid(row=3,columnspan=3)

    def returnBook(self):
    
        output_file_name = 'books.xlsx'
        wb = load_workbook(output_file_name, data_only=True)
        ws = wb['BooksOut']

        # go through rows until you find a matching row, then delete it
        for row in ws.iter_rows(min_row=2):
            if row[0].value == int(self.bookIdValue.get()) and row[2].value == self.studentNameValue.get():
                temp = []
                for cell in row:
                    temp.append(cell.value)
                temp.extend(['Return', datetime.now()])
                print(temp)
                ws.delete_rows(row[0].row)

                # add the return to the log
                ws2 = wb['Log']
                ws2.append(temp)
                wb.save(output_file_name)
App()